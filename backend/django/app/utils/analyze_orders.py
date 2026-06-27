import re
import sys
import csv
from datetime import datetime

# ── Patterns ──────────────────────────────────────────────────────────────────
ts_fmt = "%Y-%m-%d %H:%M:%S,%f"

price_diff_pat = re.compile(
    r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d+).*Price diff for \S+:.*?"
    r"'ask_diff': ([0-9.]+).*?'bid_diff': ([0-9.]+).*?"
    r"'primary_ask': ([0-9.]+).*?'hedge_ask': ([0-9.]+).*?"
    r"'primary_bid': ([0-9.]+).*?'hedge_bid': ([0-9.]+)"
)
filled_pat = re.compile(
    r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d+).*\[UserDataStream\] Order [A-Z_]+: "
    r"side=(\w+) fill_price=([0-9.]+) avg_price=[0-9.]+ qty=([0-9.]+)/[0-9.]+ order_id=(\d+)"
)
# Order placement (the moment the order_id was first opened on the book)
placed_pat = re.compile(
    r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d+).*Chase order placed: "
    r"order_id=(\d+) side=(\w+) qty=([0-9.]+) price=([0-9.]+)"
)
# MT5 new hedge (action=1): price is request[5]
hedge_new_pat = re.compile(
    r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d+).*Order successful:.*?"
    r"'request': \[1, \d+, \d+, '\S+', ([0-9.]+), ([0-9.]+),"
)
# Ground-truth primary position from the exchange (signed amount; + long, - short).
primary_pos_pat = re.compile(
    r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d+).*Primary Position \S+ - Amount: (-?[0-9.]+)"
)


# ── Functions ─────────────────────────────────────────────────────────────────

# Timestamp columns rendered as Excel times (hh:mm:ss.000).
DATE_COLS = ('fill_ts', 'placed_ts', 'pred_ts', 'hedge_ts')
DATE_FMT  = 'hh:mm:ss.000'
# Price/price-diff columns rendered to 2 decimal places.
PRICE_COLS = (
    'pred_primary_price', 'pred_hedge_price', 'predicted_price_diff',
    'actual_primary_price', 'actual_hedge_price', 'actual_price_diff',
    'primary_price_diff', 'hedge_price_diff', 'vwap',
)
PRICE_FMT = '0.00'
# Columns that get a zero-axis data-bar conditional format.
DATA_BAR_COLS = ('slippage', 'primary_edge', 'hedge_edge', 'net_edge')

# Human-readable, multi-line column headers (newlines wrap in the cell so the
# column can stay narrow). Falls back to the raw field name if not listed.
HEADERS = {
    'fill_ts'              : 'Fill\nTime',
    'side'                 : 'Side',
    'vol'                  : 'Vol',
    'cumvol'               : 'Cum\nVol',
    'primary_pos'          : 'Primary\nPos',
    'order_id'            : 'Order ID',
    'placed_ts'           : 'Placed\nTime',
    'pred_ts'             : 'Pred\nTime',
    'pred_to_fill_ms'     : 'Pred→Fill\n(ms)',
    'order_age_ms'        : 'Order Age\n(ms)',
    'pred_to_hedge_ms'    : 'Pred→Hedge\n(ms)',
    'pred_primary_price'  : 'Pred\nPrimary',
    'pred_hedge_price'    : 'Pred\nHedge',
    'predicted_price_diff': 'Pred\nDiff',
    'hedge_ts'            : 'Hedge\nTime',
    'actual_primary_price': 'Actual\nPrimary',
    'actual_hedge_price'  : 'Actual\nHedge',
    'actual_price_diff'   : 'Actual\nDiff',
    'primary_price_diff'  : 'Primary\nDiff',
    'hedge_price_diff'    : 'Hedge\nDiff',
    'slippage'            : 'Slippage',
    'primary_edge'        : 'Primary\nEdge',
    'hedge_edge'          : 'Hedge\nEdge',
    'net_edge'            : 'Net\nEdge',
    'vwap'                : 'VWAP\nSpread',
    'pnl'                 : 'PnL',
    'cumpnl'              : 'Cum\nPnL',
}


def _write_xlsx(out_xlsx, fieldnames, rows):
    """Write matched orders to a styled .xlsx workbook.

    - first row frozen, human-readable multi-line headers
    - timestamp columns formatted as hh:mm:ss.000
    - column widths fit their content
    - zero-axis data bars on slippage / primary_edge / hedge_edge / net_edge
      (positive values to the right in green, negative to the left in red)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'orders'

    headers = [HEADERS.get(fn, fn) for fn in fieldnames]
    ws.append(headers)
    header_align = Alignment(wrap_text=True, horizontal='center', vertical='center')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = header_align
    # Two lines tall so wrapped headers are fully visible.
    ws.row_dimensions[1].height = 30

    for r in rows:
        out_row = []
        for fn in fieldnames:
            v = r.get(fn, '')
            if v == '' or v is None:
                out_row.append(None)
            elif fn in DATE_COLS:
                out_row.append(datetime.strptime(v, '%Y-%m-%d %H:%M:%S.%f'))
            else:
                out_row.append(v)
        ws.append(out_row)

    ws.freeze_panes = 'A2'
    last_row = ws.max_row

    databar_ranges = []
    for idx, fn in enumerate(fieldnames, 1):
        col = get_column_letter(idx)

        if fn in DATE_COLS:
            for row_i in range(2, last_row + 1):
                ws[f'{col}{row_i}'].number_format = DATE_FMT
        elif fn in PRICE_COLS:
            for row_i in range(2, last_row + 1):
                ws[f'{col}{row_i}'].number_format = PRICE_FMT

        # Width to fit content. Header contributes only its longest line, so a
        # wrapped multi-line header lets the column stay narrow.
        width = max(len(line) for line in HEADERS.get(fn, fn).split('\n'))
        for row_i in range(2, last_row + 1):
            cell = ws[f'{col}{row_i}']
            if cell.value is None:
                disp = ''
            elif fn in DATE_COLS:
                disp = '00:00:00.000'
            elif fn in PRICE_COLS:
                disp = f'{cell.value:.2f}'
            else:
                disp = str(cell.value)
            width = max(width, len(disp))
        # Data-bar columns get extra width so the bars on both sides of the
        # centred zero axis have room to read.
        pad = 8 if fn in DATA_BAR_COLS else 2
        ws.column_dimensions[col].width = width + pad

        if fn in DATA_BAR_COLS and last_row >= 2:
            # Centre every bar so the zero axis sits in the middle of the
            # column, with positive bars to the right and negative to the left.
            databar_ranges.append((f'{col}2:{col}{last_row}', 'middle'))

    wb.save(out_xlsx)
    if databar_ranges:
        _inject_databars(out_xlsx, databar_ranges)


def _inject_databars(path, ranges,
                     positive='FF63C384', negative='FFFF0000', axis='FF000000'):
    """Add Excel-style data bars with a zero axis to ``ranges`` in ``path``.

    openpyxl's ``DataBarRule`` only writes the 2007 data bar, which fills every
    bar from the left edge (min→max) with one colour. Excel's default data bars
    use the ``x14`` extension: a zero axis with positive values drawn to the
    right and negative values to the left in a separate colour. openpyxl can't
    emit that, so we patch the worksheet XML after saving.

    ``ranges`` is a list of ``(sqref, axis_position)`` pairs, where
    ``axis_position`` is ``'automatic'`` (axis placed from the data range) or
    ``'middle'`` (zero axis fixed at the centre of the column).
    """
    import os
    import uuid
    import zipfile

    sheet = 'xl/worksheets/sheet1.xml'
    X14_NS = 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/main'
    XM_NS = 'http://schemas.microsoft.com/office/excel/2006/main'

    with zipfile.ZipFile(path) as z:
        xml = z.read(sheet).decode('utf-8')

    legacy_blocks, x14_rules = [], []
    for rng, axis_pos in ranges:
        guid = '{%s}' % str(uuid.uuid4()).upper()
        # 2007 data bar (older Excel / LibreOffice fallback). It points at the
        # x14 rule below via a shared GUID.
        legacy_blocks.append(
            f'<conditionalFormatting sqref="{rng}">'
            f'<cfRule type="dataBar" priority="1">'
            f'<dataBar><cfvo type="min"/><cfvo type="max"/>'
            f'<color rgb="{positive}"/></dataBar>'
            f'<extLst><ext xmlns:x14="{X14_NS}" '
            f'uri="{{B025F937-C7B1-47D3-B67F-A62EFF666E3E}}">'
            f'<x14:id>{guid}</x14:id></ext></extLst>'
            f'</cfRule></conditionalFormatting>'
        )
        # x14 data bar: zero axis + distinct negative colour.
        x14_rules.append(
            f'<x14:conditionalFormatting xmlns:xm="{XM_NS}">'
            f'<x14:cfRule type="dataBar" id="{guid}">'
            f'<x14:dataBar minLength="0" maxLength="100" '
            f'negativeBarColorSameAsPositive="0" '
            f'negativeBarBorderColorSameAsPositive="0" axisPosition="{axis_pos}">'
            f'<x14:cfvo type="autoMin"/><x14:cfvo type="autoMax"/>'
            f'<x14:negativeFillColor rgb="{negative}"/>'
            f'<x14:axisColor rgb="{axis}"/>'
            f'</x14:dataBar></x14:cfRule>'
            f'<xm:sqref>{rng}</xm:sqref>'
            f'</x14:conditionalFormatting>'
        )

    ext_xml = (
        '<extLst><ext uri="{78C0D931-6437-407d-A8EE-F0AAD7539E65}" '
        f'xmlns:x14="{X14_NS}"><x14:conditionalFormattings>'
        + ''.join(x14_rules) +
        '</x14:conditionalFormattings></ext></extLst>'
    )

    # conditionalFormatting must come before <pageMargins; the worksheet
    # extLst must be the final child of <worksheet>.
    xml = xml.replace('<pageMargins', ''.join(legacy_blocks) + '<pageMargins', 1)
    xml = xml.replace('</worksheet>', ext_xml + '</worksheet>', 1)

    tmp = path + '.tmp'
    with zipfile.ZipFile(path) as zsrc, \
            zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zdst:
        for item in zsrc.infolist():
            data = xml.encode('utf-8') if item.filename == sheet else zsrc.read(item.filename)
            zdst.writestr(item, data)
    os.replace(tmp, path)


def _book_realized_pnl(rows, initial_pos=0.0):
    """Assign per-fill ``pnl`` / running ``cumpnl`` / ``cumvol`` to ``rows``.

    ``rows`` are all fills (matched and unmatched) in chronological order.
    ``initial_pos`` is the signed primary position the book already carried
    before the first fill (from the exchange's Primary Position log), so cumvol
    tracks the real position rather than assuming it starts flat. Its spread
    basis is unknown, so it is seeded at the first observed spread.

    The primary position is tracked in signed lots (BUY = long +, SELL = short -).
    Each fill carries a spread basis ``d = actual_price_diff`` (primary - hedge),
    and the open position keeps a volume-weighted average basis ``vwap``.

    Realized PnL is booked *only* on the volume that closes against an existing
    opposite-side position:
        closing a long  (an incoming SELL): (d - vwap) * closed_vol
        closing a short (an incoming BUY) : (vwap - d) * closed_vol
    A same-side fill just grows the position and blends the VWAP; it books 0.
    If a fill is larger than the open position it flips the side: the closing
    part realizes PnL, and the remainder opens a fresh position at this fill's
    spread ``d``.

    A fill with no hedge leg has no ``actual_price_diff``; its spread is unknown,
    so it is marked at the last observed spread (``mark``). Its volume still
    flows through the position so ``cumvol`` stays correct.

    Returns the final ``(pos, vwap, mark)`` so the caller can mark open inventory.
    """
    pos = float(initial_pos)   # signed primary lots; + = net long primary
    # Seed the basis of any inherited position and the mark at the first spread
    # we observe, so closing inherited inventory books sane (near-zero) PnL.
    first_spread = next((r['actual_price_diff'] for r in rows
                         if r.get('actual_price_diff') is not None), 0.0)
    vwap = first_spread if pos else 0.0   # avg spread basis of the open position
    mark = first_spread                   # last observed spread (hedge-less proxy)
    cum = 0.0
    for r in rows:
        vol = r['vol']
        signed = vol if r['side'] == 'BUY' else -vol

        if r.get('correction'):
            # Synthetic resync to the exchange's logged position: adjust the
            # volume only. The missing trades' basis is unknown so book no
            # realized PnL; blend the open basis at the last mark when growing,
            # leave it unchanged when shrinking, reset it on a sign flip.
            new_pos = pos + signed
            if new_pos == 0:
                vwap = 0.0
            elif pos == 0 or (pos > 0) == (new_pos > 0):
                if abs(new_pos) > abs(pos):
                    vwap = (vwap * abs(pos) + mark * (abs(new_pos) - abs(pos))) / abs(new_pos)
            else:
                vwap = mark
            pos = new_pos
            r['pnl'] = 0.0
            r['cumpnl'] = cum
            r['cumvol'] = round(pos, 4)
            r['vwap'] = round(vwap, 4)
            continue

        d = r.get('actual_price_diff')
        eff = d if d is not None else mark   # hedge-less fill -> mark at last spread

        realized = 0.0
        if pos == 0 or (pos > 0) == (signed > 0):
            # Flat or same direction: grow the position, blend the VWAP.
            new_pos = pos + signed
            vwap = (vwap * abs(pos) + eff * vol) / abs(new_pos)
            pos = new_pos
        else:
            # Opposite direction: realize PnL on the closed volume.
            close_vol = min(vol, abs(pos))
            realized = (eff - vwap) * close_vol if pos > 0 else (vwap - eff) * close_vol
            pos += signed
            if pos == 0:
                vwap = 0.0
            elif (pos > 0) == (signed > 0):
                # Flipped past flat: remainder opens a new position at eff.
                vwap = eff
            # else: position only reduced, VWAP unchanged.

        if d is not None:
            mark = d

        r['pnl'] = round(realized, 4)
        cum = round(cum + realized, 4)
        r['cumpnl'] = cum
        r['cumvol'] = round(pos, 4)   # signed net primary position after this fill
        r['vwap'] = round(vwap, 4)    # avg spread basis of the open position
    return pos, vwap, mark


def _sync_positions(fill_rows, primary_positions, initial_pos):
    """Interleave correction rows so cumvol stays in sync with the exchange.

    ``fill_rows`` must be in chronological order. For each fill the actual
    primary position is taken from the last ``Primary Position`` log that falls
    after that fill and before the next one, and stored on the row as
    ``primary_pos``. Whenever the running position (seeded from ``initial_pos``)
    drifts from that logged amount -- e.g. fills that never reached this log --
    a synthetic ``correction`` row carrying the difference is inserted so the
    position re-syncs to reality.

    Returns the augmented row list (fills + correction rows, in order).
    """
    for r in fill_rows:
        r.setdefault('primary_pos', None)
    if not fill_rows or not primary_positions:
        return list(fill_rows)

    ts = [datetime.strptime(r['fill_ts'], '%Y-%m-%d %H:%M:%S.%f') for r in fill_rows]
    n = len(fill_rows)
    augmented = []
    running = float(initial_pos)
    for i, r in enumerate(fill_rows):
        t0 = ts[i]
        t1 = ts[i + 1] if i + 1 < n else datetime.max
        # Last logged position in this fill's post-fill window.
        snaps = [p['amount'] for p in primary_positions if t0 <= p['ts_dt'] < t1]
        snap = snaps[-1] if snaps else None
        r['primary_pos'] = snap

        running += r['vol'] if r['side'] == 'BUY' else -r['vol']
        augmented.append(r)

        if snap is not None:
            disc = round(snap - running, 4)
            if abs(disc) > 1e-6:
                augmented.append({
                    'fill_ts'     : r['fill_ts'],
                    'side'        : 'BUY' if disc > 0 else 'SELL',
                    'vol'         : round(abs(disc), 4),
                    'order_id'    : 'SYNC',
                    'primary_pos' : snap,
                    'correction'  : True,
                    'match'       : False,
                })
                running = snap
    return augmented


def analyze_orders(log_file, out_xlsx=None):
    import os
    base = os.path.basename(log_file)
    dir_ = os.path.dirname(os.path.abspath(log_file))
    if out_xlsx is None:
        out_xlsx = os.path.join(dir_, f"{base}_orders.xlsx")
    else:
        # Always emit a real .xlsx regardless of the extension passed in.
        out_xlsx = os.path.splitext(out_xlsx)[0] + ".xlsx"
    out_log = os.path.join(dir_, f"{base}_filtered.log")

    price_diffs = []
    fills = []
    hedges_new = []
    placements = {}
    primary_positions = []
    filtered_lines = []

    with open(log_file, 'r', encoding='utf-8', errors='replace') as f:
        for line in f:
            m = price_diff_pat.search(line)
            if m:
                price_diffs.append({
                    'ts_str': m.group(1).replace(',', '.'), 'ts_dt': datetime.strptime(m.group(1), ts_fmt),
                    'ask_diff': float(m.group(2)), 'bid_diff': float(m.group(3)),
                    'primary_ask': float(m.group(4)), 'hedge_ask': float(m.group(5)),
                    'primary_bid': float(m.group(6)), 'hedge_bid': float(m.group(7)),
                })
                filtered_lines.append((price_diffs[-1]['ts_dt'], line))
                continue
            m = filled_pat.search(line)
            if m:
                fills.append({
                    'ts_str': m.group(1).replace(',', '.'), 'ts_dt': datetime.strptime(m.group(1), ts_fmt),
                    'side': m.group(2), 'fill_price': float(m.group(3)), 'vol': float(m.group(4)), 'order_id': m.group(5)
                })
                filtered_lines.append((fills[-1]['ts_dt'], line))
                continue
            m = hedge_new_pat.search(line)
            if m:
                hedges_new.append({
                    'ts_str': m.group(1).replace(',', '.'), 'ts_dt': datetime.strptime(m.group(1), ts_fmt),
                    'volume': float(m.group(2)), 'price': float(m.group(3)),
                    'used': False
                })
                filtered_lines.append((hedges_new[-1]['ts_dt'], line))
                continue
            m = placed_pat.search(line)
            if m:
                order_id = m.group(2)
                # Keep the first placement seen for this order_id.
                if order_id not in placements:
                    placements[order_id] = {
                        'ts_str': m.group(1).replace(',', '.'), 'ts_dt': datetime.strptime(m.group(1), ts_fmt),
                        'side': m.group(3), 'vol': float(m.group(4)), 'price': float(m.group(5)),
                    }
                filtered_lines.append((datetime.strptime(m.group(1), ts_fmt), line))
                continue
            m = primary_pos_pat.search(line)
            if m:
                primary_positions.append({
                    'ts_dt': datetime.strptime(m.group(1), ts_fmt),
                    'amount': float(m.group(2)),
                })
                filtered_lines.append((primary_positions[-1]['ts_dt'], line))

    filtered_lines.sort(key=lambda x: x[0])
    with open(out_log, 'w', encoding='utf-8') as f:
        f.writelines(line for _, line in filtered_lines)
    print(f"Filtered log  : {out_log}")

    print(f"Price diffs   : {len(price_diffs)}")
    print(f"Entry fills   : {len(fills)}")
    print(f"MT5 new hedges: {len(hedges_new)}")
    print(f"Order placed  : {len(placements)}")
    print(f"Primary pos   : {len(primary_positions)}")

    # Ground-truth starting position: the last logged primary position at or
    # before the first fill. cumvol is seeded with this so it tracks the real
    # exchange position instead of assuming the book starts flat.
    initial_pos = 0.0
    if fills and primary_positions:
        first_fill_ts = min(f['ts_dt'] for f in fills)
        prior = [p for p in primary_positions if p['ts_dt'] <= first_fill_ts]
        if prior:
            initial_pos = prior[-1]['amount']
    if initial_pos:
        print(f"Initial pos   : {initial_pos:+g}  (cumvol start, from Primary Position log)")
    print()

    results = []

    for fill in fills:
        ft = fill['ts_dt']

        # The fill is reported asynchronously via the user data stream, so the
        # price diff logged just before the FILLED line is unrelated to this
        # order. Track back via order_id to when the order was first placed,
        # and use the price diff that was current at that moment instead.
        placed = placements.get(fill['order_id'])
        ref_ts = placed['ts_dt'] if placed else ft

        # Last price diff strictly before the order was placed (or before the
        # fill if no placement was found, e.g. order placed outside this log).
        pred = None
        for pd in reversed(price_diffs):
            if pd['ts_dt'] <= ref_ts:
                pred = pd
                break

        # First unused MT5 new hedge after fill (within 15 s)
        hedge = None
        for h in hedges_new:
            if h['used']:
                continue
            delta = (h['ts_dt'] - ft).total_seconds()
            if 0 <= delta <= 15:
                hedge = h
                h['used'] = True
                break

        if pred is None or hedge is None:
            # Incomplete row (missing the price-diff prediction and/or the
            # hedge fill), so the edge/slippage analytics can't be computed.
            # It is still a real fill that moves the position, so emit a flat
            # row with whatever we have. When the hedge leg exists the spread
            # basis is known, so the fill fully participates in PnL; otherwise
            # only its volume flows through (see _book_realized_pnl).
            row = {
                'fill_ts'              : fill['ts_str'],
                'side'                 : fill['side'],
                'vol'                  : fill['vol'],
                'order_id'             : fill['order_id'],
                'placed_ts'            : placed['ts_str'] if placed else '',
                'actual_primary_price' : fill['fill_price'],
                'match'                : False,
            }
            if hedge is not None:
                row['hedge_ts']           = hedge['ts_str']
                row['actual_hedge_price'] = hedge['price']
                row['actual_price_diff']  = round(fill['fill_price'] - hedge['price'], 4)
            results.append(row)
            continue

        is_buy = fill['side'] == 'BUY'
        pred_primary_price     = pred['primary_bid'] if is_buy else pred['primary_ask']
        pred_hedge_price      = pred['hedge_bid'] if is_buy else pred['hedge_ask']
        predicted_price_diff  = pred['bid_diff']  if is_buy else pred['ask_diff']

        actual_primary_price = fill['fill_price']
        actual_hedge_price   = hedge['price']
        actual_price_diff    = round(actual_primary_price - actual_hedge_price, 4)
        primary_price_diff   = round(actual_primary_price - pred_primary_price, 4)
        hedge_price_diff     = round(actual_hedge_price - pred_hedge_price, 4)
        slippage             = round(actual_price_diff - predicted_price_diff, 4)

        # Re-sign vs. the order side so positive always means "better than
        # predicted" and negative always means "worse than predicted":
        #   SELL: higher primary fill / lower hedge fill is better
        #   BUY:  lower primary fill / higher hedge fill is better
        sign         = 1 if fill['side'] == 'SELL' else -1
        primary_edge = round(sign * primary_price_diff, 4)
        hedge_edge   = round(-sign * hedge_price_diff, 4)
        net_edge     = round(sign * slippage, 4)

        # pnl / cumpnl are filled in afterwards by a position-tracking pass
        # (realized only on volume that closes against the opposite side).

        pred_to_fill_ms      = round((fill['ts_dt'] - pred['ts_dt']).total_seconds() * 1000)
        pred_to_hedge_ms     = round((hedge['ts_dt'] - pred['ts_dt']).total_seconds() * 1000)
        order_age_ms         = round((fill['ts_dt'] - placed['ts_dt']).total_seconds() * 1000) if placed else ''

        results.append({
            'fill_ts'              : fill['ts_str'],
            'side'                 : fill['side'],
            'vol'                  : fill['vol'],
            'order_id'             : fill['order_id'],
            'placed_ts'            : placed['ts_str'] if placed else '',
            'pred_ts'              : pred['ts_str'],
            'pred_to_fill_ms'      : pred_to_fill_ms,
            'order_age_ms'         : order_age_ms,
            'pred_to_hedge_ms'     : pred_to_hedge_ms,
            'pred_primary_price'   : pred_primary_price,
            'pred_hedge_price'     : pred_hedge_price,
            'predicted_price_diff' : predicted_price_diff,
            'hedge_ts'             : hedge['ts_str'],
            'actual_primary_price' : actual_primary_price,
            'actual_hedge_price'   : actual_hedge_price,
            'actual_price_diff'    : actual_price_diff,
            'primary_price_diff'   : primary_price_diff,
            'hedge_price_diff'     : hedge_price_diff,
            'slippage'             : slippage,
            'primary_edge'         : primary_edge,
            'hedge_edge'           : hedge_edge,
            'net_edge'             : net_edge,
            'match'                : True
        })

    # Every fill (matched or not) moves the real position. Sync the running
    # position to the exchange's logged Primary Position (interleaving any
    # correction rows), then run the position-tracking pass over all rows in
    # chronological order. It books realized PnL and the running
    # cumvol/cumpnl/vwap onto each row in place.
    fill_rows = sorted(results, key=lambda x: x['fill_ts'])
    sheet_rows = _sync_positions(fill_rows, primary_positions, initial_pos)
    pos, vwap, mark = _book_realized_pnl(sheet_rows, initial_pos)

    matched     = [r for r in sheet_rows if r.get('match')]
    corrections = [r for r in sheet_rows if r.get('correction')]
    unmatched   = [r for r in sheet_rows if not r.get('match') and not r.get('correction')]

    fieldnames = [
        'fill_ts', 'side', 'vol', 'cumvol', 'primary_pos', 'order_id', 'placed_ts', 'pred_ts',
        'pred_to_fill_ms', 'order_age_ms', 'pred_to_hedge_ms',
        'pred_primary_price', 'pred_hedge_price', 'predicted_price_diff',
        'hedge_ts', 'actual_primary_price', 'actual_hedge_price', 'actual_price_diff',
        'primary_price_diff', 'hedge_price_diff', 'slippage',
        'primary_edge', 'hedge_edge', 'net_edge', 'vwap', 'pnl', 'cumpnl',
    ]
    _write_xlsx(out_xlsx, fieldnames, sheet_rows)
    print(f"Written to {out_xlsx}")
    print()

    col = (
        f"{'#':<3} {'Fill Time':<26} {'S':<5} {'Vol':>8} {'CumVol':>8} {'Order ID':<14}"
        f"{'P→Fill':>8} {'OrderAge':>10} {'P→Hedge':>8} "
        f"{'PredPrimary':>11} {'PredHedge':>10} {'PredDiff':>9}"
        f"{'ActPrimary':>11} {'ActHedge':>10} {'ActDiff':>9}"
        f"{'PrimaryDiff':>12} {'HedgeDiff':>10} {'Slippage':>9}"
        f"{'PrimaryEdge':>12} {'HedgeEdge':>10} {'NetEdge':>9}"
        f"{'Pnl':>10} {'CumPnl':>10}"
    )
    print(col)
    print("-" * len(col))

    for i, r in enumerate(matched, 1):
        s = f"{r['slippage']:+.4f}"
        age = f"{r['order_age_ms']}ms" if r['order_age_ms'] != '' else '-'
        print(
            f"{i:<3} {r['fill_ts']:<26} {r['side']:<5} {r['vol']:>8} {r['cumvol']:>+8.3f} {r['order_id']:<14}"
            f"{r['pred_to_fill_ms']:>7}ms {age:>10} {r['pred_to_hedge_ms']:>7}ms"
            f"{r['pred_primary_price']:>11.2f} {r['pred_hedge_price']:>10.2f} {r['predicted_price_diff']:>9.2f}"
            f"{r['actual_primary_price']:>11.2f} {r['actual_hedge_price']:>10.2f} {r['actual_price_diff']:>9.4f}"
            f"{r['primary_price_diff']:>+12.4f} {r['hedge_price_diff']:>+10.4f} {s:>9}"
            f"{r['primary_edge']:>+12.4f} {r['hedge_edge']:>+10.4f} {r['net_edge']:>+9.4f}"
            f"{r['pnl']:>+10.4f} {r['cumpnl']:>+10.4f}"
        )

    print()
    if matched:
        slippages = [r['slippage'] for r in matched]
        sell_slip = [r['slippage'] for r in matched if r['side'] == 'SELL']
        buy_slip  = [r['slippage'] for r in matched if r['side'] == 'BUY']

        sep = "-" * 55
        print(sep)
        print(f"  Matched orders : {len(matched)}")
        print(f"  Avg slippage   : {sum(slippages)/len(slippages):+.4f}")
        print(f"  Min slippage   : {min(slippages):+.4f}")
        print(f"  Max slippage   : {max(slippages):+.4f}")
        if sell_slip:
            print(f"  SELL avg slip  : {sum(sell_slip)/len(sell_slip):+.4f}  "
                  f"(n={len(sell_slip)}, range [{min(sell_slip):+.4f} to {max(sell_slip):+.4f}])")
        if buy_slip:
            print(f"  BUY  avg slip  : {sum(buy_slip)/len(buy_slip):+.4f}  "
                  f"(n={len(buy_slip)}, range [{min(buy_slip):+.4f} to {max(buy_slip):+.4f}])")
        worse  = sum(1 for r in matched if r['net_edge'] < 0)
        better = sum(1 for r in matched if r['net_edge'] > 0)
        print(f"  Worse than pred: {worse}  |  Better: {better}")
        print(sep)

        # ── PnL (in price-points x lots; multiply by point value for money) ──
        # Realized is booked fill-by-fill over all fills (matched and unmatched)
        # in time order: a fill only books PnL on the volume that closes against
        # the opposite-side position, valued against that position's running
        # VWAP spread (see _book_realized_pnl). The leftover inventory is open
        # and marked at the last observed spread.
        realized = sheet_rows[-1]['cumpnl']

        if pos > 0:            # net long primary, close by selling at mark
            unrealized = (mark - vwap) * pos
            open_desc  = f"long primary {pos:.3f} @ spread {vwap:+.4f}"
        elif pos < 0:          # net short primary, close by buying at mark
            unrealized = (vwap - mark) * (-pos)
            open_desc  = f"short primary {-pos:.3f} @ spread {vwap:+.4f}"
        else:
            unrealized, open_desc = 0.0, "flat"

        print("  PnL")
        print(f"  Realized       : {realized:+.4f}")
        print(f"  Open inventory : {open_desc}")
        print(f"  Unrealized     : {unrealized:+.4f}  (marked @ last spread {mark:+.4f})")
        print(f"  MTM total      : {realized + unrealized:+.4f}")
        print(f"  Unmatched fills: {len(unmatched)}")
        if corrections:
            net_corr = sum(c['vol'] if c['side'] == 'BUY' else -c['vol'] for c in corrections)
            print(f"  Sync corrections: {len(corrections)}  (net {net_corr:+g} lots to match logged position)")
        print(sep)


def analyze_chase_delay(log_file, out_csv=None, symbol='XAUUSDT', threshold=0.10):
    if out_csv is None:
        import os
        base = os.path.basename(log_file)
        out_csv = os.path.join(os.path.dirname(os.path.abspath(log_file)), f"{base}_chase_delay.csv")
    chase_re = re.compile(
        r'(\d{2}:\d{2}:\d{2},\d{3}).*Chase order placed: order_id=(\d+) side=(\w+) qty=([\d.]+) price=([\d.]+)'
    )
    pdiff_re = re.compile(
        rf'(\d{{2}}:\d{{2}}:\d{{2}},\d{{3}}).*Price diff for {symbol}.*primary_ask.: ([\d.]+).*primary_bid.: ([\d.]+)'
    )

    chases = []
    pdiffs = []

    with open(log_file, encoding='utf-8', errors='replace') as f:
        for line in f:
            m = chase_re.search(line)
            if m:
                chases.append({'time': m.group(1), 'order_id': m.group(2), 'side': m.group(3),
                                'vol': m.group(4), 'price': float(m.group(5))})
            m = pdiff_re.search(line)
            if m:
                pdiffs.append({'time': m.group(1), 'primary_ask': float(m.group(2)), 'primary_bid': float(m.group(3))})

    def time_to_ms(t):
        h, m, s = t.split(':')
        s, ms = s.split(',')
        return int(h) * 3600000 + int(m) * 60000 + int(s) * 1000 + int(ms)

    rows = []
    for c in chases:
        ct = time_to_ms(c['time'])
        p = c['price']

        best = None
        for pd in pdiffs:
            ref = pd['primary_ask'] if c['side'] == 'BUY' else pd['primary_bid']
            if abs(ref - p) <= threshold:
                if best is None or abs(time_to_ms(pd['time']) - ct) < abs(time_to_ms(best['time']) - ct):
                    best = pd

        if best:
            delay_ms = time_to_ms(best['time']) - ct
            matched = best['primary_ask'] if c['side'] == 'BUY' else best['primary_bid']
            price_seen_before = delay_ms <= 0
        else:
            delay_ms = None
            matched = None
            price_seen_before = None

        rows.append({
            'order_id':           c['order_id'],
            'chase_time':         c['time'],
            'side':               c['side'],
            'vol':                c['vol'],
            'chase_price':        p,
            'matched_pdiff_time': best['time'] if best else '',
            'delay_ms':           delay_ms if delay_ms is not None else '',
            'matched_price':      matched if matched is not None else '',
            'price_seen_before':  ('YES' if price_seen_before else 'NO') if price_seen_before is not None else '',
        })

    with open(out_csv, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=[
            'order_id', 'chase_time', 'side', 'vol', 'chase_price',
            'matched_pdiff_time', 'delay_ms', 'matched_price', 'price_seen_before',
        ])
        writer.writeheader()
        writer.writerows(rows)

    seen_before = sum(1 for r in rows if r['price_seen_before'] == 'YES')
    not_seen    = sum(1 for r in rows if r['price_seen_before'] == 'NO')
    no_match    = sum(1 for r in rows if r['price_seen_before'] == '')
    delayed     = sum(1 for r in rows if isinstance(r['delay_ms'], int) and r['delay_ms'] > 500)

    print(f"Chase orders analysed : {len(rows)}")
    print(f"Price seen before      : {seen_before}  (delay <= 0)")
    print(f"Price seen after       : {not_seen}  (delay > 0)")
    print(f"  of which > 500ms    : {delayed}")
    print(f"No match in log        : {no_match}")
    print(f"Written to {out_csv}")


# ── Dispatch ──────────────────────────────────────────────────────────────────

COMMANDS = {
    '--chase-delay': {
        'usage': '<log_file> [out_csv]',
        'nargs': 1,
        'fn': lambda args: analyze_chase_delay(*args),
    },
}

if len(sys.argv) >= 2 and sys.argv[1] in COMMANDS:
    cmd = COMMANDS[sys.argv[1]]
    if len(sys.argv) < 2 + cmd['nargs']:
        print(f"Usage: python {sys.argv[0]} {sys.argv[1]} {cmd['usage']}")
        sys.exit(1)
    cmd['fn'](sys.argv[2:])
    sys.exit(0)

if len(sys.argv) < 2:
    print(f"Usage: python {sys.argv[0]} <log_file> [out_csv]")
    print(f"       python {sys.argv[0]} --chase-delay <log_file> [out_csv]")
    sys.exit(1)

analyze_orders(*sys.argv[1:])
